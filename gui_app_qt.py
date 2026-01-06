import sys
import os
import shutil
import subprocess
import openpyxl
import time as time_module
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QProgressBar, QTextEdit, QMessageBox, QCheckBox,
                             QSpinBox, QGroupBox, QGridLayout, QSplashScreen)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QPixmap, QColor, QPainter, QFont
from check_tax_official import check_cccd_official, setup_driver


class LoadingSplash(QSplashScreen):
    """Custom splash screen with loading progress."""
    
    def __init__(self):
        # Create a pixmap for the splash screen
        pixmap = QPixmap(400, 250)
        pixmap.fill(QColor("#FFFFFF"))  # White background
        
        super().__init__(pixmap)
        self.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.FramelessWindowHint)
        
        # Add progress bar
        self.progress = QProgressBar(self)
        self.progress.setGeometry(50, 200, 300, 20)
        self.progress.setTextVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                background-color: #f5f5f5;
            }
            QProgressBar::chunk {
                background-color: #b84626;
                border-radius: 3px;
            }
        """)
        self.progress.setValue(0)
        self._message = "Loading..."
        
    def drawContents(self, painter: QPainter):
        """Draw custom splash content."""
        # Title
        painter.setPen(QColor("#b84626"))  # Talentnet accent color
        title_font = QFont("Arial", 22, QFont.Weight.Bold)
        painter.setFont(title_font)
        painter.drawText(self.rect().adjusted(0, 50, 0, 0), 
                        Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop, 
                        "Talentnet TaxTracker")
        
        # Loading message
        msg_font = QFont("Arial", 11)
        painter.setFont(msg_font)
        painter.setPen(QColor("#666666"))
        painter.drawText(self.rect().adjusted(0, 100, 0, 0), 
                        Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop, 
                        self._message)
        
        # Version/info
        info_font = QFont("Arial", 9)
        painter.setFont(info_font)
        painter.setPen(QColor("#999999"))
        painter.drawText(self.rect().adjusted(0, 130, 0, 0), 
                        Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop, 
                        "Initializing application...")
        
    def setProgress(self, value: int, message: str = ""):
        """Update progress and message."""
        self._message = message
        self.progress.setValue(value)
        self.repaint()
        QApplication.processEvents()


class WorkerThread(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int, str)
    timing_signal = pyqtSignal(float, float, float)  # total_time, avg_time, est_remaining
    finished_signal = pyqtSignal()
    
    def __init__(self, input_path, open_browser=False, get_max_retries=None, get_delay_seconds=None, start_index=0, output_folder=None):
        super().__init__()
        self.input_path = input_path
        self.open_browser = open_browser
        self.output_folder = output_folder  # User-selected output folder
        # Use getter functions to read config dynamically
        self.get_max_retries = get_max_retries or (lambda: 20)
        self.get_delay_seconds = get_delay_seconds or (lambda: 2)
        self.start_index = start_index
        self.is_running = True
        self.is_paused = False
        self.current_index = 0

    def log_wrapper(self, message):
        self.log_signal.emit(message)

    def run(self):
        driver = None  # Shared browser instance
        try:
            # Create output filename with timestamp in selected output folder
            output_dir = self.output_folder or os.path.dirname(self.input_path)
            base_name = os.path.basename(self.input_path)
            name, ext = os.path.splitext(base_name)
            timestamp = datetime.now().strftime("%d%m%y%H%M")
            output_path = os.path.join(output_dir, f"{name}_processed_{timestamp}{ext}")
            
            # Copy file first (only if not resuming)
            if self.start_index == 0:
                shutil.copy2(self.input_path, output_path)
                self.log_signal.emit(f"Created working copy: {output_path}")
            else:
                self.log_signal.emit(f"Resuming from index {self.start_index}")
            
            # Load workbook
            wb = openpyxl.load_workbook(output_path)
            sheet = wb.active
            
            # Map columns
            headers = {}
            for cell in sheet[1]:
                if cell.value:
                    headers[str(cell.value).strip()] = cell.column
            
            required_cols = ["CMND/CCCD", "MST 1", "Tên người nộp thuế", "Cơ quan thuế", "Ghi chú MST 1"]
            missing = [c for c in required_cols if c not in headers]
            
            if missing:
                self.log_signal.emit(f"Error: Missing columns: {', '.join(missing)}")
                self.finished_signal.emit()
                return

            col_map = {col: headers[col] for col in required_cols}
            
            # Find rows to process
            rows_to_process = []
            max_row = sheet.max_row
            
            for row_idx in range(2, max_row + 1):
                mst_val = sheet.cell(row=row_idx, column=col_map["MST 1"]).value
                cccd_val = sheet.cell(row=row_idx, column=col_map["CMND/CCCD"]).value
                
                if cccd_val and not mst_val:
                    rows_to_process.append(row_idx)
            
            total = len(rows_to_process)
            self.log_signal.emit(f"Found {total} rows to process")
            
            # Create browser ONCE at the start (reuse for all checks)
            self.log_signal.emit("Starting browser (will be reused for all checks)...")
            driver = setup_driver(self.open_browser, self.log_wrapper)
            
            # Timing tracking
            batch_start_time = time_module.time()
            processed_count = 0
            total_processing_time = 0
            
            for i, row_idx in enumerate(rows_to_process):
                self.current_index = i
                
                # Skip if resuming
                if i < self.start_index:
                    continue
                
                if not self.is_running:
                    self.log_signal.emit(f"Stopped at index {i}")
                    break
                
                # Handle pause
                while self.is_paused and self.is_running:
                    self.msleep(100)
                    
                cccd_val = sheet.cell(row=row_idx, column=col_map["CMND/CCCD"]).value
                cccd_str = str(cccd_val).strip()
                
                progress_pct = int(((i) / total) * 100)
                self.progress_signal.emit(progress_pct, f"Processing {i+1}/{total}: {cccd_str}")
                
                try:
                    # Read current config values (allows dynamic updates)
                    current_max_retries = self.get_max_retries()
                    current_delay = self.get_delay_seconds()
                    
                    # Track time for this record
                    record_start_time = time_module.time()
                    
                    # Pass the shared driver instance
                    result = check_cccd_official(
                        cccd_str, 
                        open_browser=self.open_browser, 
                        log_callback=self.log_wrapper,
                        max_retries=current_max_retries,
                        delay_seconds=current_delay,
                        driver=driver  # Reuse the same browser
                    )
                    
                    # Calculate timing stats
                    record_time = time_module.time() - record_start_time
                    processed_count += 1
                    total_processing_time += record_time
                    
                    # Calculate stats
                    total_elapsed = time_module.time() - batch_start_time
                    avg_time = total_processing_time / processed_count if processed_count > 0 else 0
                    remaining_records = total - (i + 1)
                    est_remaining = avg_time * remaining_records
                    
                    # Emit timing signal
                    self.timing_signal.emit(total_elapsed, avg_time, est_remaining)
                    
                    # Update cells
                    if result.get("tax_id"):
                        sheet.cell(row=row_idx, column=col_map["MST 1"]).value = result["tax_id"]
                    if result.get("name"):
                        sheet.cell(row=row_idx, column=col_map["Tên người nộp thuế"]).value = result["name"]
                    if result.get("place"):
                        sheet.cell(row=row_idx, column=col_map["Cơ quan thuế"]).value = result["place"]
                    if result.get("status"):
                        sheet.cell(row=row_idx, column=col_map["Ghi chú MST 1"]).value = result["status"]
                    
                    # Save immediately
                    wb.save(output_path)
                    self.log_signal.emit(f"Processed {cccd_str}: {result.get('status')} - {result.get('tax_id')}")
                    
                except Exception as e:
                    self.log_signal.emit(f"Error processing {cccd_str}: {str(e)}")
            
            if self.is_running:
                self.progress_signal.emit(100, "Completed")
                self.log_signal.emit("Processing complete. File saved.")
            
        except Exception as e:
            self.log_signal.emit(f"Critical Error: {str(e)}")
        finally:
            # Close the shared browser at the end
            if driver:
                self.log_signal.emit("Closing browser...")
                try:
                    driver.quit()
                except Exception:
                    pass
            self.finished_signal.emit()

    def stop(self):
        self.is_running = False
        
    def pause(self):
        self.is_paused = True
        
    def resume(self):
        self.is_paused = False

class TaxCheckerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Talentnet TaxTracker")
        self.setGeometry(100, 100, 800, 650)
        
        self.file_path = ""
        self.output_path = ""
        self.output_folder = ""  # Output folder path
        self.worker = None
        self.last_stopped_index = 0
        self.log_file = None  # Current log file path
        
        # Create log folder
        self.log_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log")
        os.makedirs(self.log_folder, exist_ok=True)
        
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # File Selection
        file_layout = QHBoxLayout()
        self.path_label = QLabel("No file selected")
        self.path_label.setStyleSheet("border: 1px solid #ccc; padding: 5px; background: white; color: #333333;")
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        
        file_layout.addWidget(self.path_label, stretch=1)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Output Folder Selection
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("Output Folder:"))
        self.output_folder_label = QLabel("Same as input file")
        self.output_folder_label.setStyleSheet("border: 1px solid #ccc; padding: 5px; background: white; color: #333333;")
        output_browse_btn = QPushButton("Browse")
        output_browse_btn.clicked.connect(self.browse_output_folder)
        self.open_folder_btn = QPushButton("📂 Open Folder")
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        self.open_folder_btn.setStyleSheet("background-color: #607D8B; color: white; padding: 5px 10px;")
        
        output_layout.addWidget(self.output_folder_label, stretch=1)
        output_layout.addWidget(output_browse_btn)
        output_layout.addWidget(self.open_folder_btn)
        layout.addLayout(output_layout)
        
        # Configuration Group
        config_group = QGroupBox("Configuration")
        config_layout = QGridLayout()
        
        # Max Retries
        config_layout.addWidget(QLabel("Max Retries:"), 0, 0)
        self.retries_spinbox = QSpinBox()
        self.retries_spinbox.setRange(1, 100)
        self.retries_spinbox.setValue(20)
        config_layout.addWidget(self.retries_spinbox, 0, 1)
        
        # Delay Seconds
        config_layout.addWidget(QLabel("Delay (seconds):"), 0, 2)
        self.delay_spinbox = QSpinBox()
        self.delay_spinbox.setRange(0, 60)
        self.delay_spinbox.setValue(2)
        config_layout.addWidget(self.delay_spinbox, 0, 3)
        
        # Show Browser checkbox
        self.browser_checkbox = QCheckBox("Show Browser")
        self.browser_checkbox.setChecked(False)
        config_layout.addWidget(self.browser_checkbox, 0, 4)
        
        config_group.setLayout(config_layout)
        layout.addWidget(config_group)
        
        # Action Buttons
        btn_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("▶ Start")
        self.start_btn.clicked.connect(self.start_processing)
        self.start_btn.setEnabled(False)
        self.start_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px;")
        
        self.pause_btn = QPushButton("⏸ Pause")
        self.pause_btn.clicked.connect(self.pause_processing)
        self.pause_btn.setEnabled(False)
        self.pause_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 8px 16px;")
        
        self.continue_btn = QPushButton("▶ Continue")
        self.continue_btn.clicked.connect(self.continue_processing)
        self.continue_btn.setEnabled(False)
        self.continue_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px 16px;")
        
        # self.stop_btn = QPushButton("⏹ Stop")
        # self.stop_btn.clicked.connect(self.stop_processing)
        # self.stop_btn.setEnabled(False)
        # self.stop_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px;")
        
        # self.restart_btn = QPushButton("🔄 Restart")
        # self.restart_btn.clicked.connect(self.restart_processing)
        # self.restart_btn.setEnabled(False)
        # self.restart_btn.setStyleSheet("background-color: #9C27B0; color: white; padding: 8px 16px;")
        
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.pause_btn)
        btn_layout.addWidget(self.continue_btn)
        # btn_layout.addWidget(self.stop_btn)
        # btn_layout.addWidget(self.restart_btn)
        
        
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Status
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("font-weight: bold; padding: 5px;")
        layout.addWidget(self.status_label)
        
        # Timing Stats
        self.timing_label = QLabel("")
        self.timing_label.setStyleSheet("color: #666666; padding: 5px; font-size: 11px;")
        layout.addWidget(self.timing_label)
        
        # Progress Bar
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)
        
        # Log Area
        layout.addWidget(QLabel("Logs:"))
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        layout.addWidget(self.log_area)
        
        # Contact Info
        contact_label = QLabel("Contact if any problem: nguyenlvc@talentnetgroup.com | phumg@talentnetgroup.com")
        contact_label.setStyleSheet("color: #666666; font-size: 11px; padding: 5px;")
        contact_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(contact_label)
        
    def browse_file(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if filename:
            self.file_path = filename
            self.path_label.setText(filename)
            self.start_btn.setEnabled(True)
            # self.restart_btn.setEnabled(True)
            
            # Set default output folder to input file's directory
            if not self.output_folder:
                self.output_folder = os.path.dirname(self.file_path)
                self.output_folder_label.setText(self.output_folder)
            
            # Compute output path with timestamp
            base_name = os.path.basename(self.file_path)
            name, ext = os.path.splitext(base_name)
            timestamp = datetime.now().strftime("%d%m%y%H%M")
            self.output_path = os.path.join(self.output_folder, f"{name}_processed_{timestamp}{ext}")
            
            # Check if output exists
            if os.path.exists(self.output_path):
                self.log(f"Found existing result file: {self.output_path}")
                
            self.log(f"Selected file: {filename}")
    
    def browse_output_folder(self):
        """Browse for output folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.output_folder or "")
        if folder:
            self.output_folder = folder
            self.output_folder_label.setText(folder)
            self.log(f"Output folder set to: {folder}")
            
            # Update output path if file is already selected
            if self.file_path:
                base_name = os.path.basename(self.file_path)
                name, ext = os.path.splitext(base_name)
                timestamp = datetime.now().strftime("%d%m%y%H%M")
                self.output_path = os.path.join(self.output_folder, f"{name}_processed_{timestamp}{ext}")
    
    def open_output_folder(self):
        """Open output folder in file explorer."""
        folder = self.output_folder or (os.path.dirname(self.file_path) if self.file_path else "")
        if folder and os.path.exists(folder):
            try:
                if sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', folder], check=False)
                elif sys.platform == 'win32':  # Windows
                    os.startfile(folder)
                else:  # Linux
                    subprocess.run(['xdg-open', folder], check=False)
                self.log(f"Opened folder: {folder}")
            except Exception as e:
                self.log(f"Could not open folder: {e}")
        else:
            self.log("No output folder selected")
            
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {message}"
        self.log_area.append(log_line)
        # Scroll to bottom
        cursor = self.log_area.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        self.log_area.setTextCursor(cursor)
        
        # Write to log file if active
        if self.log_file:
            try:
                with open(self.log_file, "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
            except Exception:
                pass  # Silently ignore log file write errors
        
    def start_processing(self):
        if not self.file_path or not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Error", "Invalid file path")
            return
        
        # Create new log file with timestamp
        log_timestamp = datetime.now().strftime("%d%m%y%H%M")
        self.log_file = os.path.join(self.log_folder, f"log_{log_timestamp}.txt")
        self.log(f"Log file created: {self.log_file}")
            
        self._start_worker(start_index=0)
        
    def continue_processing(self):
        if self.worker and self.worker.is_paused:
            # Resume paused worker
            self.worker.resume()
            self.log("Resumed processing")
            self.pause_btn.setEnabled(True)
            self.continue_btn.setEnabled(False)
            self.status_label.setText("Processing...")
        elif self.last_stopped_index > 0:
            # Start new worker from last stopped index
            self._start_worker(start_index=self.last_stopped_index)
            
    def pause_processing(self):
        if self.worker:
            self.worker.pause()
            self.log("Paused processing")
            self.pause_btn.setEnabled(False)
            self.continue_btn.setEnabled(True)
            self.status_label.setText("Paused")
            
    def stop_processing(self):
        if self.worker:
            self.last_stopped_index = self.worker.current_index
            self.worker.stop()
            self.log(f"Stopping... (can continue from index {self.last_stopped_index})")
            
    def restart_processing(self):
        if self.worker:
            self.worker.stop()
            self.worker.wait()
        self.last_stopped_index = 0
        self._start_worker(start_index=0)
        
    def _start_worker(self, start_index=0):
        self.start_btn.setEnabled(False)
        self.pause_btn.setEnabled(True)
        self.continue_btn.setEnabled(False)
        # self.stop_btn.setEnabled(True)
        # self.restart_btn.setEnabled(True)
        
        # Compute output path with timestamp using selected output folder
        output_dir = self.output_folder or os.path.dirname(self.file_path)
        base_name = os.path.basename(self.file_path)
        name, ext = os.path.splitext(base_name)
        timestamp = datetime.now().strftime("%d%m%y%H%M")
        self.output_path = os.path.join(output_dir, f"{name}_processed_{timestamp}{ext}")
        
        self.log(f"Starting processing from index {start_index}...")
        self.progress_bar.setValue(0)
        self.status_label.setText("Processing...")
        self.timing_label.setText("")
        
        open_browser = self.browser_checkbox.isChecked()
        
        self.worker = WorkerThread(
            self.file_path, 
            open_browser=open_browser,
            get_max_retries=lambda: self.retries_spinbox.value(),
            get_delay_seconds=lambda: self.delay_spinbox.value(),
            start_index=start_index,
            output_folder=self.output_folder
        )
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.timing_signal.connect(self.update_timing)
        self.worker.finished_signal.connect(self.processing_finished)
        self.worker.start()
        
    def update_progress(self, value, status):
        self.progress_bar.setValue(value)
        self.status_label.setText(status)
    
    def update_timing(self, total_elapsed, avg_time, est_remaining):
        """Update the timing statistics label."""
        def format_time(seconds):
            """Format seconds into human-readable time."""
            if seconds < 60:
                return f"{seconds:.1f}s"
            elif seconds < 3600:
                mins = int(seconds // 60)
                secs = int(seconds % 60)
                return f"{mins}m {secs}s"
            else:
                hours = int(seconds // 3600)
                mins = int((seconds % 3600) // 60)
                return f"{hours}h {mins}m"
        
        timing_text = (
            f"⏱ Elapsed: {format_time(total_elapsed)}  |  "
            f"Avg/record: {format_time(avg_time)}  |  "
            f"ETA: {format_time(est_remaining)}"
        )
        self.timing_label.setText(timing_text)
        
    def processing_finished(self):
        self.start_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.continue_btn.setEnabled(self.last_stopped_index > 0)
        # self.stop_btn.setEnabled(False)
        # self.restart_btn.setEnabled(True)
        self.status_label.setText("Finished")
        self.worker = None
        
        # Close log file
        if self.log_file:
            self.log(f"Processing finished. Log saved to: {self.log_file}")
            self.log_file = None
        
    def open_result(self):
        if self.output_path and os.path.exists(self.output_path):
            try:
                if sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', self.output_path], check=False)
                elif sys.platform == 'win32':  # Windows
                    os.startfile(self.output_path)
                else:  # Linux
                    subprocess.run(['xdg-open', self.output_path], check=False)
                self.log(f"Opened: {self.output_path}")
            except Exception as e:
                self.log(f"Could not open file: {e}")
                print(f"Could not open file: {e}")
        else:
            self.log("Result file not found")
            print("Result file not found")
    
    def closeEvent(self, event):
        """Handle window close - stop worker thread and quit completely."""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(2000)  # Wait up to 2 seconds for thread to finish
            if self.worker.isRunning():
                self.worker.terminate()  # Force terminate if still running
        
        # Ensure application quits completely
        QApplication.quit()
        event.accept()


def main():
    """Main entry point with splash screen."""
    app = QApplication(sys.argv)
    
    # Create and show splash screen
    splash = LoadingSplash()
    splash.show()
    QApplication.processEvents()
    
    # Simulate loading steps with progress updates
    splash.setProgress(30, "Loading components...")
    QApplication.processEvents()
    
    splash.setProgress(60, "Initializing UI...")
    QApplication.processEvents()
    
    # Create main window
    splash.setProgress(90, "Starting application...")
    QApplication.processEvents()
    
    window = TaxCheckerApp()
    
    # Brief delay to show the splash screen (minimum 500ms visibility)
    QTimer.singleShot(500, lambda: finish_loading(splash, window))
    
    sys.exit(app.exec())


def finish_loading(splash, window):
    """Close splash and show main window."""
    splash.setProgress(100, "Ready!")
    window.show()
    splash.finish(window)


if __name__ == "__main__":
    main()
