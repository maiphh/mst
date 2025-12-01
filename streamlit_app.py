import streamlit as st
import pandas as pd
import openpyxl
import os
import time
from check_tax_official import check_cccd_official
from io import BytesIO

st.set_page_config(page_title="Tax Checker Tool", layout="wide")

st.title("Tax Checker Tool")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

if "processing" not in st.session_state:
    st.session_state.processing = False

if "logs" not in st.session_state:
    st.session_state.logs = []

def log(message):
    st.session_state.logs.append(message)

if uploaded_file:
    st.write("File uploaded successfully.")
    
    # Load workbook to check headers
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column
        
        required_cols = ["CMND/CCCD", "MST 1", "Tên người nộp thuế", "Cơ quan thuế", "Ghi chú MST 1"]
        missing_cols = [col for col in required_cols if col not in headers]
        
        if missing_cols:
            st.error(f"Missing columns: {', '.join(missing_cols)}")
        else:
            col_map = {col: headers[col] for col in required_cols}
            
            if st.button("Start Processing", disabled=st.session_state.processing):
                st.session_state.processing = True
                st.session_state.logs = []
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                log_container = st.container()
                
                # Create a copy of the workbook for processing
                # We need to save the uploaded file to a temporary path or keep in memory
                # openpyxl can work with file-like objects, but for saving we might want a BytesIO
                
                # We will process row by row
                max_row = sheet.max_row
                rows_to_process = []
                
                # Identify rows to process
                for row_idx in range(2, max_row + 1):
                    mst_val = sheet.cell(row=row_idx, column=col_map["MST 1"]).value
                    cccd_val = sheet.cell(row=row_idx, column=col_map["CMND/CCCD"]).value
                    
                    if cccd_val and not mst_val:
                        rows_to_process.append(row_idx)
                
                total_rows = len(rows_to_process)
                st.write(f"Found {total_rows} rows to process.")
                
                for i, row_idx in enumerate(rows_to_process):
                    cccd_val = sheet.cell(row=row_idx, column=col_map["CMND/CCCD"]).value
                    cccd_str = str(cccd_val).strip()
                    
                    status_text.text(f"Processing {i+1}/{total_rows}: CCCD {cccd_str}")
                    
                    try:
                        result = check_cccd_official(cccd_str)
                        
                        # Update cells
                        if result.get("tax_id"):
                            sheet.cell(row=row_idx, column=col_map["MST 1"]).value = result["tax_id"]
                        if result.get("name"):
                            sheet.cell(row=row_idx, column=col_map["Tên người nộp thuế"]).value = result["name"]
                        if result.get("place"):
                            sheet.cell(row=row_idx, column=col_map["Cơ quan thuế"]).value = result["place"]
                        if result.get("status"):
                            sheet.cell(row=row_idx, column=col_map["Ghi chú MST 1"]).value = result["status"]
                            
                        log(f"Row {row_idx} ({cccd_str}): {result.get('status')} - {result.get('tax_id')}")
                        
                    except Exception as e:
                        log(f"Error Row {row_idx} ({cccd_str}): {e}")
                    
                    # Update progress
                    progress_bar.progress((i + 1) / total_rows)
                    
                    # Update logs display
                    with log_container:
                        st.text("\n".join(st.session_state.logs[-5:])) # Show last 5 logs
                
                st.session_state.processing = False
                status_text.text("Processing complete!")
                
                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="Download Processed Excel",
                    data=output,
                    file_name="processed_tax_check.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
    except Exception as e:
        st.error(f"Error loading file: {e}")

# Display full logs at the bottom
with st.expander("Full Logs"):
    for l in st.session_state.logs:
        st.text(l)
